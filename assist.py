#!/bin/python3

import os.path

import click
import pandas as pd
import questionary
from openpyxl import load_workbook


class ContestChecker:
    def __init__(self):
        self.standings = None
        self.grades = None
        self.grades_xlsx = None
        self.contests = {}
        self.cur_contest = ""
        self.losers_count = 0
        self.full_grade_count = 0
        self.manual_check = []

    def run(self):
        self.input_fields()
        self.work_table()

    def find_filepath(self, substring):
        results = []
        for path in os.listdir("."):
            if substring in path:
                results.append(path)
        return results

    def work_sheet(self, sheet):
        print(f"⚙️ Processing {self.cur_contest} - {sheet.title}")

        for row in range(1, sheet.max_row):
            student_name = sheet.cell(row, 1).value
            if student_name is None:
                continue
            if student_name == "Student":
                continue

            student_login = sheet.cell(row, 1).value
            student_login = student_login.strip() if student_login else "None"

            student_row = self.standings[
                (self.standings["user_name"] == str(student_name))
                | (self.standings["login"] == str(student_login))
            ]

            # print(sheet.title, row)
            if student_row.shape[0] == 0:
                print(
                    f"    ⚠️ Student {student_name} [{student_login}] is not found. Putting zeros..."
                )
                self.losers_count += 1
                for ind, problem in enumerate(student_row.iloc[:, 3:-1]):
                    sheet.cell(row, self.contests[self.cur_contest] + ind).value = 0
                continue

            for ind, problem in enumerate(student_row.iloc[:, 3:-1]):
                mark = int(is_solved(student_row[problem].iloc[0]))
                if problem.split("(")[0] not in self.manual_check:
                    sheet.cell(row, self.contests[self.cur_contest] + ind).value = mark
                elif mark == 0:
                    sheet.cell(row, self.contests[self.cur_contest] + ind).value = 0
                else:
                    continue

    def input_fields(self):
        standings_csv = questionary.select(
            "Select the standings file", choices=self.find_filepath("standings")
        ).ask()  # returns value of selection

        self.grades_xlsx = questionary.select(
            "Select the grades file", choices=self.find_filepath("xlsx")
        ).ask()  # returns value of selection

        self.grades = load_workbook(self.grades_xlsx)

        first_sheet = self.grades.worksheets[1]
        for col in range(2, first_sheet.max_column):
            col_name = first_sheet.cell(1, col).value
            if col_name is not None and "Contest" in col_name:
                self.contests[col_name] = col

        print(self.contests)
        self.cur_contest = questionary.select(
            "Select the contest to check",
            choices=list(reversed([x for x in self.contests])),
        ).ask()

        self.standings = pd.read_csv(standings_csv)

        self.req_string = questionary.text("Enter the requirements string").ask()

        if self.req_string:
            self.archive_file = questionary.select(
                "Select the submits file",
                choices=self.find_filepath("zip"),
            ).ask()  # returns value of selection
            self.checker = LLMChecker(self.req_string, self.archive_file)
            self.manual_check = list(self.checker.requirements.keys())

        print(f"✅ Done reading csv: {self.standings.shape[0]} participants.")
        print(
            f'✅ Mean score: {self.standings["Score"].mean().round(2)}/{self.standings["Score"].max()}'
        )

    def work_table(self):
        for sheet in self.grades.worksheets:
            if sheet.title == "All" or "2" not in sheet.title:
                continue
            self.work_sheet(sheet)

        self.grades.save(self.grades_xlsx)
        print(f"\n✅ {self.cur_contest} was checked successfully!")
        full_grades_count = sum(
            (self.standings["Score"] == self.standings.shape[0] - 4)
        )
        print(f"⚠️ {self.losers_count} students did not do their homework :(")
        print(f"🎉 {full_grades_count} students have a 100% grade!")


class QuestionaryOption(click.Option):
    def __init__(self, param_decls=None, **attrs):
        click.Option.__init__(self, param_decls, **attrs)
        if not isinstance(self.type, click.Choice):
            raise Exception("ChoiceOption type arg must be click.Choice")

    def prompt_for_value(self, ctx):
        val = questionary.select(self.prompt, choices=self.type.choices).unsafe_ask()
        return val


def is_solved(problem_standing):
    if not problem_standing:
        return 0

    mark_symbol = str(problem_standing)[0]

    if mark_symbol.isdigit():
        if "+" in problem_standing:
            return 1
        return int(mark_symbol) > 0
    else:
        return 0


if __name__ == "__main__":
    checker = ContestChecker()
    checker.run()
